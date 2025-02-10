# app/routes/reports.py
from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file, make_response
from app.services.report_service import ReportService
from app.config.settings import Settings
import io
from io import BytesIO
import zipfile
import json
from openpyxl.cell.cell import MergedCell
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
from app import db
from sqlalchemy import or_
from app.models.schedule import Schedule
from concurrent.futures import ThreadPoolExecutor, as_completed

reports = Blueprint('reports', __name__)


@reports.route('/reports')
def index():
    """Главная страница отчетов"""
    return render_template('reports/index.html')


from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file, make_response
from app.services.report_service import ReportService
from app.config.settings import Settings
import io
from io import BytesIO
import zipfile
import json
from openpyxl.cell.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

reports = Blueprint('reports', __name__)


@reports.route('/reports')
def index():
    """Главная страница отчетов"""
    return render_template('reports/index.html')


@reports.route('/reports/teacher-load')
def teacher_load():
    """Отчет по нагрузке преподавателя"""
    teacher_name = request.args.get('teacher')
    semester = request.args.get('semester', Settings.get_current_semester(), type=int)
    week = request.args.get('week', type=int)

    if not teacher_name:
        # Показываем список всех преподавателей с их общей нагрузкой
        teachers_load = ReportService.get_teachers_total_load(semester)
        return render_template('reports/teachers_list.html',
                               teachers_load=teachers_load,
                               semester=semester)

    report = ReportService.get_teacher_load(teacher_name, semester, week)
    return render_template('reports/teacher_load.html', report=report)


@reports.route('/reports/teacher-load/export')
def export_teacher_load():
    """Экспорт нагрузки одного преподавателя"""
    teacher_name = request.args.get('teacher')
    semester = request.args.get('semester', type=int)

    if not teacher_name or not semester:
        return jsonify({'error': 'Отсутствуют обязательные параметры'}), 400

    try:
        excel_file = ReportService.export_teacher_load_excel(teacher_name, semester)

        filename = f'Нагрузка_{teacher_name}_{semester}сем.xlsx'
        filename = filename.encode('utf-8').decode('latin-1')

        response = make_response(excel_file.getvalue())
        response.headers.update({
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Length': len(excel_file.getvalue())
        })

        return response

    except Exception as e:
        print(f"Ошибка экспорта: {str(e)}")
        return jsonify({'error': str(e)}), 500


@reports.route('/reports/teacher-load/export-multiple', methods=['GET', 'POST'])
def export_multiple_teachers():
    try:
        if request.method == 'POST':
            data = request.get_json()
            export_type = data.get('type')
            semester = data.get('semester')
            teachers = data.get('teachers', [])
            sort_by_faculty = data.get('sort_by_faculty', '0') == '1'
        else:
            export_type = request.args.get('type')
            semester = request.args.get('semester', type=int)
            teachers = json.loads(request.args.get('teachers', '[]'))
            sort_by_faculty = request.args.get('sort_by_faculty', '0') == '1'

        if not teachers or not semester:
            return jsonify({'error': 'Отсутствуют обязательные параметры'}), 400

        # Получаем все данные для преподавателей одним запросом
        all_reports = ReportService.get_multiple_teachers_load(semester, teachers)

        # Функция для генерации одного Excel-файла
        def generate_excel_for_teacher(teacher, faculty=None):
            try:
                teacher_report = all_reports[teacher]
                if faculty:
                    # Фильтруем по факультету
                    filtered_subjects = {
                        subj: data for subj, data in teacher_report['subjects'].items()
                        if any(g['faculty'] == faculty for g in data['groups'].values())
                    }
                    filtered_report = {
                        'teacher_name': teacher_report['teacher_name'],
                        'semester': teacher_report['semester'],
                        'subjects': filtered_subjects
                    }
                    excel_file = ReportService.export_teacher_load_excel_from_report(filtered_report)
                else:
                    excel_file = ReportService.export_teacher_load_excel_from_report(teacher_report)

                filename = f'Нагрузка_{teacher}_{semester}сем.xlsx'
                return filename, excel_file.getvalue()
            except Exception as e:
                print(f"Ошибка при обработке преподавателя {teacher}: {str(e)}")
                error_wb = Workbook()
                error_ws = error_wb.active
                error_ws['A1'] = f"Ошибка при создании отчета для {teacher}: {str(e)}"
                error_file = BytesIO()
                error_wb.save(error_file)
                return f'Ошибка_{teacher}_{semester}сем.xlsx', error_file.getvalue()

        memory_file = BytesIO()

        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as main_zip:
            if sort_by_faculty:
                # Сортируем преподавателей по факультетам
                faculties = {}
                for teacher in teachers:
                    teacher_report = all_reports.get(teacher)
                    if not teacher_report:
                        continue
                    teacher_faculties = set()
                    for subj, subj_data in teacher_report['subjects'].items():
                        for grp_data in subj_data['groups'].values():
                            teacher_faculties.add(grp_data['faculty'])

                    for faculty in teacher_faculties:
                        faculties.setdefault(faculty, []).append(teacher)

                # Обработка по факультетам с помощью пула потоков
                for faculty, faculty_teachers in faculties.items():
                    faculty_zip = BytesIO()
                    with zipfile.ZipFile(faculty_zip, 'w', zipfile.ZIP_DEFLATED) as sub_zip:
                        # Используем ThreadPoolExecutor для параллельной генерации
                        with ThreadPoolExecutor(max_workers=4) as executor:
                            futures = [executor.submit(generate_excel_for_teacher, t, faculty) for t in
                                       faculty_teachers]

                            for future in as_completed(futures):
                                fname, fdata = future.result()
                                sub_zip.writestr(fname, fdata)

                    faculty_zip.seek(0)
                    main_zip.writestr(f'{faculty}.zip', faculty_zip.getvalue())
            else:
                # Обработка всех преподавателей в один архив без сортировки по факультетам
                with ThreadPoolExecutor(max_workers=4) as executor:
                    futures = [executor.submit(generate_excel_for_teacher, t) for t in teachers]

                    for future in as_completed(futures):
                        fname, fdata = future.result()
                        main_zip.writestr(fname, fdata)

        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'Нагрузка_преподавателей_{semester}сем.zip'
        )
    except Exception as e:
        print(f"Ошибка экспорта: {str(e)}")
        return jsonify({'error': str(e)}), 500


@reports.route('/reports/teacher-load/export-list')
def export_teacher_load_list():
    """Экспорт расписания преподавателей в формате списка"""
    try:
        export_type = request.args.get('type')
        semester = request.args.get('semester', type=int)
        teachers = json.loads(request.args.get('teachers', '[]'))
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not all([teachers, semester, start_date, end_date]):
            return jsonify({'error': 'Отсутствуют обязательные параметры'}), 400

        if export_type == 'zip':
            memory_file = BytesIO()
            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                for teacher in teachers:
                    excel_file = ReportService.generate_schedule_list(teacher, start_date, end_date)
                    filename = f'Расписание_{teacher}_{start_date}_{end_date}.xlsx'
                    filename = filename.encode('utf-8').decode('latin-1')
                    zf.writestr(filename, excel_file.getvalue())

            memory_file.seek(0)
            return send_file(
                memory_file,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f'Расписание_преподавателей_{start_date}_{end_date}.zip'
            )
        else:
            # If only one teacher or single file requested
            teacher = teachers[0] if len(teachers) == 1 else teachers[0]  # Take first teacher if multiple selected
            excel_file = ReportService.generate_schedule_list(teacher, start_date, end_date)

            filename = f'Расписание_{teacher}_{start_date}_{end_date}.xlsx'
            filename = filename.encode('utf-8').decode('latin-1')

            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

    except Exception as e:
        print(f"Ошибка экспорта: {str(e)}")
        return jsonify({'error': str(e)}), 500

    @reports.errorhandler(Exception)
    def handle_error(error):
        """Обработчик ошибок для роутов отчетов"""
        print(f"Error in reports route: {str(error)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': str(error)}), 500
        flash(f'Произошла ошибка: {str(error)}', 'error')
        return redirect(url_for('reports.index'))

    # Также добавим логирование ошибок
    import logging
    logger = logging.getLogger(__name__)

    def log_error(error):
        """Логирование ошибок с дополнительной информацией"""
        logger.error(f"""
        Error occurred:
        Type: {type(error).__name__}
        Message: {str(error)}
        Route: {request.url}
        Method: {request.method}
        Args: {dict(request.args)}
        Data: {request.get_data(as_text=True)}
        """)


@reports.route('/reports/attendance')
def attendance():
    """Отчет по проведению занятий"""
    available_semesters = Schedule.get_available_semesters()
    semester = request.args.get('semester', type=int)
    group = request.args.get('group')
    subject = request.args.get('subject')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if 'get_groups' in request.args:
            # Получаем группы для семестра
            groups = Schedule.query.with_entities(Schedule.group_name) \
                .filter(Schedule.semester == semester) \
                .distinct() \
                .order_by(Schedule.group_name) \
                .all()
            return jsonify([g[0] for g in groups])

        elif 'get_subjects' in request.args and group:
            # Получаем предметы для группы
            subjects = Schedule.query.with_entities(Schedule.subject) \
                .filter(Schedule.semester == semester,
                        Schedule.group_name == group) \
                .distinct() \
                .order_by(Schedule.subject) \
                .all()
            return jsonify([s[0] for s in subjects])

        elif 'get_teachers' in request.args and group and subject:
            # Получаем преподавателей для группы и предмета
            teachers = Schedule.query.with_entities(Schedule.teacher_name) \
                .filter(Schedule.semester == semester,
                        Schedule.group_name == group,
                        Schedule.subject == subject) \
                .distinct() \
                .order_by(Schedule.teacher_name) \
                .all()
            return jsonify([t[0] for t in teachers])

        elif 'get_lesson_types' in request.args and group and subject:
            # Получаем типы занятий для группы и предмета
            types = Schedule.query.with_entities(Schedule.lesson_type) \
                .filter(Schedule.semester == semester,
                        Schedule.group_name == group,
                        Schedule.subject == subject) \
                .distinct() \
                .order_by(Schedule.lesson_type) \
                .all()
            return jsonify([t[0] for t in types])

        elif 'get_subgroups' in request.args and group and subject:
            # Получаем подгруппы для группы и предмета
            subgroups = Schedule.query.with_entities(Schedule.subgroup) \
                .filter(Schedule.semester == semester,
                        Schedule.group_name == group,
                        Schedule.subject == subject,
                        Schedule.subgroup != 0) \
                .distinct() \
                .order_by(Schedule.subgroup) \
                .all()
            return jsonify([s[0] for s in subgroups])

        # Получение данных с учетом всех фильтров
        teachers = request.args.getlist('teachers[]')
        lesson_types = request.args.getlist('lesson_types[]')
        subgroups = request.args.getlist('subgroups[]')

        query = Schedule.query.filter(Schedule.semester == semester)

        if group:
            query = query.filter(Schedule.group_name == group)

        if subject:
            cleaned_subject = subject.strip("'")
            query = query.filter(Schedule.subject == cleaned_subject)

        if teachers:
            query = query.filter(Schedule.teacher_name.in_(teachers))

        if lesson_types:
            query = query.filter(Schedule.lesson_type.in_(lesson_types))

        if subgroups:
            subgroups_int = [int(s) for s in subgroups]
            if 0 not in subgroups_int:  # Если не выбрана опция "Без подгруппы"
                query = query.filter(Schedule.subgroup.in_(subgroups_int))
            else:  # Если выбрана опция "Без подгруппы"
                query = query.filter(
                    or_(
                        Schedule.subgroup.in_(subgroups_int),
                        Schedule.subgroup == 0
                    )
                )

        lessons = query.order_by(
            Schedule.date,
            Schedule.time_start,
            Schedule.subgroup
        ).all()

        return render_template(
            'reports/attendance_table.html',
            lessons=lessons
        )

    # Отображение основной страницы
    return render_template(
        'reports/attendance.html',
        available_semesters=available_semesters
    )


@reports.route('/reports/attendance/export')
def attendance_export():
    """Экспорт отчета по проведению занятий"""
    try:
        semester = request.args.get('semester', type=int)
        group = request.args.get('group')
        subject = request.args.get('subject')
        teachers = request.args.getlist('teachers[]')
        lesson_types = request.args.getlist('lesson_types[]')
        subgroups = request.args.getlist('subgroups[]')

        query = Schedule.query.filter(Schedule.semester == semester)

        if group:
            query = query.filter(Schedule.group_name == group)

        if subject:
            cleaned_subject = subject.strip("'")
            query = query.filter(Schedule.subject == cleaned_subject)

        if teachers:
            query = query.filter(Schedule.teacher_name.in_(teachers))

        if lesson_types:
            query = query.filter(Schedule.lesson_type.in_(lesson_types))

        if subgroups:
            subgroups_int = [int(s) for s in subgroups]
            if 0 not in subgroups_int:
                query = query.filter(Schedule.subgroup.in_(subgroups_int))
            else:
                query = query.filter(
                    or_(
                        Schedule.subgroup.in_(subgroups_int),
                        Schedule.subgroup == 0
                    )
                )

        lessons = query.order_by(
            Schedule.date,
            Schedule.time_start,
            Schedule.subgroup
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Проведение занятий"

        # Стили
        header_font = Font(name='Times New Roman', bold=True)
        normal_font = Font(name='Times New Roman')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки колонок
        headers = ['Дата', 'Время', 'Тип занятия', 'Преподаватель', 'Аудитория', 'Подгруппа']
        for col, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header_text
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Данные
        for row, lesson in enumerate(lessons, 2):
            data = [
                lesson.date.strftime('%d.%m.%Y'),
                f"{lesson.time_start} - {lesson.time_end}",
                lesson.lesson_type,
                lesson.teacher_name,
                lesson.auditory or '',
                f"Подгруппа {lesson.subgroup}" if lesson.subgroup != 0 else ""
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='left' if col > 2 else 'center'
                )

        # Настройка ширины колонок
        column_widths = [12, 15, 15, 40, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Настройки печати
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = '1:1'

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Занятия_{semester}сем.xlsx'
        )

    except Exception as e:
        print(f"Ошибка экспорта: {str(e)}")
        return jsonify({'error': str(e)}), 500








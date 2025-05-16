# app/routes/main.py
from datetime import datetime, timedelta

from flask import render_template, request, Blueprint, jsonify, make_response, send_file
from sqlalchemy.sql.expression import func
import re
from app import db
from app.config.settings import Settings
from app.models.schedule import Schedule
from app.models.user import User
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

main = Blueprint('main', __name__, static_folder='static')


@main.route('/robots.txt')
def robots():
    content = '''
User-agent: *
Disallow: /
'''
    response = make_response(content)
    response.headers['Content-Type'] = 'text/plain'
    return response


@main.route('/')
def index():
    # Получаем данные из базы данных
    groups = db.session.query(Schedule.group_name).distinct().order_by(Schedule.group_name).all()
    teachers = db.session.query(Schedule.teacher_name).filter(Schedule.teacher_name != '').distinct().order_by(
        Schedule.teacher_name).all()
    rooms = db.session.query(Schedule.auditory).filter(Schedule.auditory != '').distinct().order_by(
        Schedule.auditory).all()

    groups = [group[0] for group in groups]
    teachers = [teacher[0] for teacher in teachers]
    rooms = [room[0] for room in rooms]

    # Определяем текущий семестр и неделю
    current_semester = Settings.get_current_semester()
    current_week = get_current_week()

    return render_template('index.html', groups=groups, teachers=teachers, rooms=rooms,
                           current_semester=current_semester, current_week=current_week)


# app/routes/main.py

@main.route('/yandex_63874f5319c0ae3b.html')
def yandex_verification():
    content = """
    <html>
        <head>
            <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
        </head>
        <body>Verification: 63874f5319c0ae3b</body>
    </html>
    """
    response = make_response(content)
    response.headers['Content-Type'] = 'text/html'
    return response


@main.route('/manifest.json')
def manifest():
    return main.send_static_file('manifest.json')


@main.route('/service-worker.js')
def service_worker():
    response = make_response(main.send_static_file('service-worker.js'))
    response.headers['Content-Type'] = 'application/javascript'
    response.headers['Cache-Control'] = 'no-cache'
    return response


@main.route('/timetable')
def schedule():
    schedule_type = request.args.get('type')
    value = request.args.get('value')

    # Определяем текущий семестр
    current_semester = Settings.get_current_semester()

    # Получаем семестр из параметров URL или используем текущий
    requested_semester = request.args.get('semester', current_semester, type=int)

    # Получаем все недели для запрошенного семестра
    weeks_data = db.session.query(
        Schedule.week_number,
        func.min(Schedule.date).label('start_date'),
        func.max(Schedule.date).label('end_date')
    ).filter_by(
        semester=requested_semester
    ).group_by(
        Schedule.week_number
    ).order_by(
        Schedule.week_number
    ).all()

    # Формируем список недель для выбранного семестра
    weeks = [week.week_number for week in weeks_data]

    # Определяем текущую/выбранную неделю
    current_week = None

    # Если неделя указана в URL, используем её
    if 'week' in request.args:
        current_week = int(request.args.get('week'))
    else:
        current_date = datetime.now().date()

        # Если смотрим текущий семестр, пытаемся найти текущую неделю
        if requested_semester == current_semester:
            for week in weeks_data:
                if week.start_date and week.end_date:
                    if week.start_date <= current_date <= week.end_date:
                        current_week = week.week_number
                        break

            # Если текущая дата не попадает ни в одну неделю,
            # находим ближайшую следующую неделю
            if not current_week and weeks_data:
                closest_future_week = None
                min_diff = timedelta.max

                for week in weeks_data:
                    if week.start_date:
                        diff = abs(week.start_date - current_date)
                        if diff < min_diff and week.start_date >= current_date:
                            min_diff = diff
                            closest_future_week = week.week_number

                if closest_future_week:
                    current_week = closest_future_week

        # Если неделя все еще не определена, берем первую доступную
        if not current_week and weeks:
            current_week = weeks[0]
        elif not current_week:
            current_week = 1

    # Получаем настройки семестров
    academic_settings = Settings.get_settings()['academic_year']
    semester_dates = {
        1: {
            'start': academic_settings['first_semester']['start'],
            'end': academic_settings['first_semester']['end']
        },
        2: {
            'start': academic_settings['second_semester']['start'],
            'end': academic_settings['second_semester']['end']
        }
    }

    today_weekday = datetime.now().isoweekday()
    settings = Settings.get_settings()

    return render_template(
        'timetable/index.html',
        current_semester=requested_semester,
        current_week=current_week,
        today_weekday=today_weekday,
        weeks=weeks,
        semester_dates=semester_dates,
        schedule_type=schedule_type,
        value=value,
        settings=settings,
        timedelta=timedelta
    )


@main.route('/api/schedule')
def get_schedule():
    schedule_type = request.args.get('type')
    value = request.args.get('value')
    semester = request.args.get('semester', type=int)
    week = request.args.get('week', type=int)

    if not all([schedule_type, value, semester, week]):
        return jsonify({'error': 'Missing required parameters'})

    try:
        # Получаем расписание
        schedule_filter = {
            'semester': semester,
            'week_number': week
        }

        if schedule_type == 'group':
            schedule_filter['group_name'] = value
        elif schedule_type == 'teacher':
            schedule_filter['teacher_name'] = value
        elif schedule_type == 'room':
            schedule_filter['auditory'] = value
        else:
            return jsonify({'error': 'Invalid schedule type'})

        schedule = Schedule.query.filter_by(**schedule_filter).all()

        # Рендерим HTML
        html = render_template(
            'timetable/schedule_table.html',
            schedule=schedule,
            schedule_type=schedule_type,
            value=value,
            timedelta=timedelta  # Передаем timedelta в шаблон
        )

        return jsonify({'html': html})

    except Exception as e:
        print(f"Error processing query: {str(e)}")
        return jsonify({'error': str(e)}), 500


def get_current_week():
    """Определяет текущую неделю семестра"""
    try:
        settings = Settings.get_settings()
        current_semester = Settings.get_current_semester()
        today = datetime.now().date()

        # Получаем все недели текущего семестра
        weeks = db.session.query(
            Schedule.week_number,
            func.min(Schedule.date).label('start_date'),
            func.max(Schedule.date).label('end_date')
        ).filter_by(
            semester=current_semester
        ).group_by(
            Schedule.week_number
        ).order_by(
            Schedule.week_number
        ).all()

        if not weeks:
            return 1

        # Ищем текущую неделю по дате
        for week in weeks:
            if week.start_date and week.end_date:
                if week.start_date <= today <= week.end_date:
                    return week.week_number

        # Если текущая дата не попадает ни в одну неделю,
        # находим ближайшую следующую неделю
        closest_future_week = None
        min_diff = timedelta.max

        for week in weeks:
            if week.start_date:
                diff = abs(week.start_date - today)
                if diff < min_diff and week.start_date >= today:
                    min_diff = diff
                    closest_future_week = week.week_number

        if closest_future_week:
            return closest_future_week

        # Если не нашли подходящую неделю, возвращаем первую доступную
        return weeks[0].week_number if weeks else 1

    except Exception as e:
        print(f"Error determining current week: {e}")
        return 1


@main.route('/init_db_2025')
def create_admin():
    try:
        # Проверяем, существует ли уже админ
        existing_admin = User.query.filter_by(is_admin=True).first()
        if existing_admin:
            return "Администратор уже существует"

        # Создаем админа
        admin = User(
            username='admin',
            email='admin@example.com',
            is_admin=True
        )
        admin.set_password('Admin123!')

        db.session.add(admin)
        db.session.commit()

        return """
        <div style='text-align: center; padding: 20px;'>
            <h2 style='color: green;'>✅ Админ создан успешно</h2>
            <p>Логин: admin</p>
            <p>Пароль: Admin123!</p>
            <p style='color: red;'>Не забудьте сменить пароль при первом входе!</p>
        </div>
        """

    except Exception as e:
        return f"Ошибка: {str(e)}"


@main.route('/free-rooms')
def free_rooms():
    """Страница поиска свободных аудиторий"""
    buildings = Schedule.get_buildings()
    time_slots = Settings.get_settings().get('time_slots', [])
    current_semester = Settings.get_current_semester()

    # Получаем максимальное количество недель из всего расписания
    max_weeks = Schedule.get_max_weeks()  # Убираем параметр semester

    return render_template(
        'free_rooms.html',
        buildings=buildings,
        time_slots=time_slots,
        current_semester=current_semester,
        current_week=Schedule.get_week_by_date(datetime.now().date(), current_semester),
        current_day=datetime.now().isoweekday(),
        weeks_count=max_weeks
    )


@main.route('/free-rooms/data')
def free_rooms_data():
    semester = request.args.get('semester', Settings.get_current_semester(), type=int)
    week = request.args.get('week', type=int)
    day = request.args.get('day', type=int)
    building = request.args.get('building')
    lesson_number = request.args.get('lesson', type=int)

    if not all([week, day, lesson_number]):
        return jsonify([])

    # Получаем время пары
    time_slots = Settings.get_settings().get('time_slots', [])
    slot_time = next((slot['start'] for slot in time_slots if slot['number'] == lesson_number), None)
    if not slot_time:
        return jsonify([])

    # Сначала получаем все аудитории текущего семестра
    base_query = db.session.query(Schedule.auditory).distinct().filter(
        Schedule.semester == semester  # Добавляем фильтр по семестру
    )

    # Применяем фильтры в зависимости от корпуса
    if building == 'other':
        base_query = base_query.filter(
            ~Schedule.auditory.regexp_match(r'^(?:[1-9]|1[0-9]|2[0-5])\.'),
            Schedule.auditory != ''
        )
    elif building == '25':
        base_query = base_query.filter(Schedule.auditory.like('25.%'))
    else:
        base_query = base_query.filter(Schedule.auditory.like(f'{building}.%'))

    # Подзапрос для получения занятых аудиторий
    busy_rooms = db.session.query(Schedule.auditory).filter(
        Schedule.semester == semester,
        Schedule.week_number == week,
        Schedule.weekday == day,
        Schedule.time_start == slot_time
    )

    # Получаем свободные аудитории, исключая занятые
    free_rooms = base_query.filter(
        ~Schedule.auditory.in_(busy_rooms)
    ).order_by(Schedule.auditory).all()

    # Преобразуем результат
    free_rooms = [room[0] for room in free_rooms if room[0]]

    return jsonify(free_rooms)


@main.route('/free-rooms/room-details')
def room_details():
    """Получение детальной информации об аудитории на день"""
    semester = request.args.get('semester', type=int)
    week = request.args.get('week', type=int)
    day = request.args.get('day', type=int)
    room = request.args.get('room')

    if not all([semester, week, day, room]):
        return jsonify({'error': 'Missing parameters'}), 400

    # Получаем временные слоты
    time_slots = Settings.get_settings().get('time_slots', [])

    # Получаем расписание для конкретной аудитории
    lessons = Schedule.query.filter(
        Schedule.semester == semester,
        Schedule.week_number == week,
        Schedule.weekday == day,
        Schedule.auditory == room
    ).order_by(
        Schedule.time_start
    ).all()

    # Создаем словарь занятости по времени
    occupied_slots = {}
    for lesson in lessons:
        start_time = lesson.time_start
        if start_time not in occupied_slots:
            occupied_slots[start_time] = []

        occupied_slots[start_time].append({
            'subject': lesson.subject,
            'teacher': lesson.teacher_name,
            'group': lesson.group_name,
            'type': lesson.lesson_type
        })

    # Формируем полное расписание
    schedule = []
    for slot in time_slots:
        start_time = slot['start']
        schedule_item = {
            'number': slot['number'],
            'time': f"{slot['start']} - {slot['end']}",
            'is_occupied': start_time in occupied_slots
        }

        if start_time in occupied_slots:
            lessons_data = occupied_slots[start_time]
            groups = [lesson['group'] for lesson in lessons_data]

            # Берем данные первого занятия, так как предмет и преподаватель обычно одинаковые
            first_lesson = lessons_data[0]
            schedule_item.update({
                'subject': first_lesson['subject'],
                'teacher': first_lesson['teacher'],
                'type': first_lesson['type'],
                'groups': groups
            })

        schedule.append(schedule_item)

    return jsonify({
        'room': room,
        'schedule': schedule
    })


@main.route('/room-comparison')
def room_comparison():
    """Page for comparing multiple rooms side by side on a specific date"""
    buildings = Schedule.get_buildings()
    time_slots = Settings.get_settings().get('time_slots', [])
    current_semester = Settings.get_current_semester()

    # Get maximum number of weeks from the Schedule model
    max_weeks = Schedule.get_max_weeks()

    # Get current date and determine weekday for default selection
    current_date = datetime.now().date()
    current_weekday = current_date.isoweekday()

    # Determine the current week based on the date
    current_week = Schedule.get_week_by_date(current_date, current_semester)

    # Get settings for the template
    settings = Settings.get_settings()

    return render_template(
        'room_comparison.html',
        buildings=buildings,
        time_slots=time_slots,
        current_semester=current_semester,
        current_week=current_week,
        current_day=current_weekday,
        weeks_count=max_weeks,
        settings=settings
    )


@main.route('/api/room-comparison')
def get_room_comparison_data():
    """API endpoint to get schedule data for multiple rooms on a specific day"""
    semester = request.args.get('semester', Settings.get_current_semester(), type=int)
    week = request.args.get('week', type=int)
    day = request.args.get('day', type=int)
    rooms = request.args.getlist('rooms[]')

    if not all([week, day]) or not rooms:
        return jsonify({'error': 'Missing required parameters'}), 400

    try:
        # Get the time slot configuration
        time_slots = Settings.get_settings().get('time_slots', [])

        # Get all schedules for the specified rooms
        rooms_data = {}

        for room in rooms:
            # Get all lessons for this room on the specified day
            lessons = Schedule.query.filter(
                Schedule.semester == semester,
                Schedule.week_number == week,
                Schedule.weekday == day,
                Schedule.auditory == room
            ).order_by(Schedule.time_start).all()

            # Organize by time slot
            room_schedule = {}

            for slot in time_slots:
                slot_key = str(slot['number'])
                slot_start = slot['start']

                # Find lessons that match this time slot
                slot_lessons = [
                    lesson for lesson in lessons
                    if lesson.time_start == slot_start
                ]

                if slot_lessons:
                    room_schedule[slot_key] = [
                        {
                            'subject': lesson.subject,
                            'teacher_name': lesson.teacher_name,
                            'group_name': lesson.group_name,
                            'lesson_type': lesson.lesson_type,
                            'subgroup': lesson.subgroup if lesson.subgroup != 0 else None,
                            'busy': True
                        } for lesson in slot_lessons
                    ]
                else:
                    room_schedule[slot_key] = [{'busy': False}]

            rooms_data[room] = room_schedule

        return jsonify({
            'success': True,
            'rooms_data': rooms_data,
            'time_slots': time_slots
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main.route('/api/search-rooms')
def search_rooms_for_comparison():
    """Search for rooms by building and search term"""
    building = request.args.get('building', '')
    search_term = request.args.get('search', '')

    try:
        # Base query for searching rooms
        query = db.session.query(Schedule.auditory).distinct().filter(Schedule.auditory != '')

        # Apply building filter
        if building == 'other':
            query = query.filter(
                ~Schedule.auditory.regexp_match(r'^(?:[1-9]|1[0-9]|2[0-5])\.'),
                Schedule.auditory != ''
            )
        elif building == '25':
            query = query.filter(Schedule.auditory.like('25.%'))
        elif building not in ['', 'all'] and building.isdigit():
            query = query.filter(Schedule.auditory.like(f'{building}.%'))

        # Apply search filter if provided
        if search_term:
            query = query.filter(Schedule.auditory.ilike(f'%{search_term}%'))

        # Get results ordered by room name
        rooms = query.order_by(Schedule.auditory).all()
        rooms = [room[0] for room in rooms if room[0]]

        return jsonify(rooms)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main.route('/api/room-comparison/export')
def export_room_comparison():
    """Endpoint для экспорта сравнения аудиторий в Excel"""
    semester = request.args.get('semester', Settings.get_current_semester(), type=int)
    week = request.args.get('week', type=int)
    day = request.args.get('day', type=int)
    rooms = request.args.getlist('rooms[]')

    if not all([week, day]) or not rooms:
        return jsonify({'error': 'Missing required parameters'}), 400

    try:
        # Получаем настройки временных слотов
        time_slots = Settings.get_settings().get('time_slots', [])
        timetable_colors = Settings.get_settings().get('appearance', {}).get('timetable_colors', {})

        # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = f"Сравнение аудиторий"

        # Назначаем стили
        header_font = Font(name='Arial', bold=True)
        normal_font = Font(name='Arial')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Добавляем заголовок
        day_names = ['', 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
        day_name = day_names[day] if 1 <= day <= 7 else f"День {day}"
        title = f"Сравнение аудиторий - {day_name}, {week} неделя, {semester} семестр"

        ws.merge_cells(f'A1:{chr(65 + len(rooms))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = header_font
        title_cell.alignment = center_align

        # Добавляем заголовки столбцов
        ws.cell(row=2, column=1, value="Время").font = header_font
        ws.cell(row=2, column=1).alignment = center_align
        ws.cell(row=2, column=1).border = border

        for i, room in enumerate(rooms):
            col = i + 2
            ws.cell(row=2, column=col, value=room).font = header_font
            ws.cell(row=2, column=col).alignment = center_align
            ws.cell(row=2, column=col).border = border

        # Получаем данные расписания и заполняем таблицу
        row_num = 3
        for slot in time_slots:
            slot_num = slot['number']
            slot_start = slot['start']
            slot_end = slot['end']

            # Добавляем ячейку с временем
            time_cell = ws.cell(row=row_num, column=1)
            time_cell.value = f"{slot_num} пара\n{slot_start} - {slot_end}"
            time_cell.alignment = center_align
            time_cell.font = normal_font
            time_cell.border = border

            # Добавляем данные для каждой аудитории
            for i, room in enumerate(rooms):
                col = i + 2

                # Получаем занятия в данной аудитории на данный временной слот
                lessons = Schedule.query.filter(
                    Schedule.semester == semester,
                    Schedule.week_number == week,
                    Schedule.weekday == day,
                    Schedule.auditory == room,
                    Schedule.time_start == slot_start
                ).all()

                cell = ws.cell(row=row_num, column=col)
                cell.border = border

                if lessons:
                    # Аудитория занята
                    cell_data = []
                    for lesson in lessons:
                        # Цвет для типа занятия
                        lesson_color = timetable_colors.get(lesson.lesson_type, "#cccccc")

                        # Формируем информацию о занятии
                        lesson_info = [
                            f"{lesson.subject} ({lesson.lesson_type})",
                            f"Преп.: {lesson.teacher_name}",
                            f"Группа: {lesson.group_name}"
                        ]
                        if lesson.subgroup != 0:
                            lesson_info[-1] += f" (Подгруппа {lesson.subgroup})"

                        cell_data.append("\n".join(lesson_info))

                    cell.value = "\n\n".join(cell_data)
                    cell.alignment = left_align

                    # Заливка фона для занятой аудитории
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                else:
                    # Аудитория свободна
                    cell.value = "Свободно"
                    cell.alignment = center_align
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            row_num += 1

        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 15
        for i in range(len(rooms)):
            ws.column_dimensions[chr(66 + i)].width = 40

        # Настройки страницы
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1

        # Конвертируем в BytesIO и отправляем пользователю
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Формируем имя файла
        filename = f"Сравнение_аудиторий_{day_name}_{week}нед_{semester}сем.xlsx"

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
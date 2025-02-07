from datetime import datetime
from typing import Dict, List
from sqlalchemy import func, distinct
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from app import db
from app.models.schedule import Schedule
from typing import List, Dict

class ReportService:
    MAX_WEEKS = 50  # Increased to handle up to week 22

    LESSON_TYPES = {
        'л.': 'lecture',
        'пр.': 'practice',
        'лаб.': 'laboratory',
        'лекция': 'lecture',
        'практика': 'practice',
        'лабораторная': 'laboratory',
        'экзамен': 'exam',
        'зач': 'test',
        'зчО': 'graded_test'
    }

    EXAM_DISPLAY = {
        'экзамен': 'Э',
        'зач.': 'З',
        'зчО.': 'ЗО'
    }

    @staticmethod
    def get_teacher_load(teacher_name: str, semester: int, week_number: int = None) -> Dict:
        """Generates a teacher load report for a given teacher and semester."""
        # Query the database for schedule entries
        query = Schedule.query.filter(Schedule.teacher_name == teacher_name, Schedule.semester == semester)
        if week_number:
            query = query.filter(Schedule.week_number == week_number)
        lessons = query.order_by(Schedule.week_number, Schedule.date).all()

        # Initialize the report dictionary with consistent week range
        report = {
            'teacher_name': teacher_name,
            'semester': semester,
            'total_hours': 0,
            'subjects': {},
            'weeks': {i: {
                'lecture': 0,
                'practice': 0,
                'laboratory': 0,
                'other': 0,
                'total': 0,
                'dates': {'start': None, 'end': None},
                'exam_info': None
            } for i in range(1, ReportService.MAX_WEEKS + 1)}  # Use MAX_WEEKS constant
        }

        # Process lessons
        for lesson in lessons:
            week_num = lesson.week_number
            subject = lesson.subject
            group = lesson.group_name
            faculty = lesson.faculty
            original_type = lesson.lesson_type
            lesson_type = ReportService.LESSON_TYPES.get(original_type, 'other')
            hours = 2  # Default hours per lesson

            # Initialize subject if not exists
            if subject not in report['subjects']:
                report['subjects'][subject] = {
                    'groups': {},
                    'total_hours': 0,
                    'by_type': {
                        'lecture': 0,
                        'practice': 0,
                        'laboratory': 0,
                        'other': 0,
                        'exam': []
                    }
                }

            # Initialize group if not exists
            if group not in report['subjects'][subject]['groups']:
                report['subjects'][subject]['groups'][group] = {
                    'faculty': faculty,
                    'total_hours': 0,
                    'by_type': {
                        'lecture': 0,
                        'practice': 0,
                        'laboratory': 0,
                        'other': 0
                    },
                    'by_week': {str(w): {
                        'lecture': 0,
                        'practice': 0,
                        'laboratory': 0,
                        'other': 0,
                        'exam_type': None,
                        'hours': {}
                    } for w in range(1, ReportService.MAX_WEEKS + 1)}  # Use MAX_WEEKS constant
                }

            group_data = report['subjects'][subject]['groups'][group]
            week_data = group_data['by_week'][str(week_num)]

            # Process lesson hours
            if lesson_type in ['lecture', 'practice', 'laboratory']:
                week_data[lesson_type] += hours
                group_data['total_hours'] += hours
                group_data['by_type'][lesson_type] += hours
                report['subjects'][subject]['total_hours'] += hours
                report['subjects'][subject]['by_type'][lesson_type] += hours

            # Add exam information if applicable
            if lesson.lesson_type in ReportService.EXAM_DISPLAY:
                week_data['exam_type'] = ReportService.EXAM_DISPLAY[lesson.lesson_type]
                report['subjects'][subject]['by_type']['exam'].append({
                    'type': ReportService.EXAM_DISPLAY[lesson.lesson_type],
                    'week': week_num,
                    'group': group
                })

        return report

    @staticmethod
    def export_teacher_load_excel(teacher_name: str, semester: int, faculty: str = None) -> BytesIO:
        """Exports teacher load to Excel with optional filtering by faculty."""
        report = ReportService.get_teacher_load(teacher_name, semester)

        if faculty:
            filtered_subjects = {
                subj: data for subj, data in report['subjects'].items()
                if any(group['faculty'] == faculty for group in data['groups'].values())
            }
            report['subjects'] = filtered_subjects

        wb = Workbook()
        ws = wb.active

        # Стили
        header_font = Font(name='Times New Roman', bold=True)
        normal_font = Font(name='Times New Roman')
        borders = {
            'thick': Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            ),
            'thin': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        center = Alignment(horizontal='center', vertical='center')
        exam_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

        # Заголовок
        ws.merge_cells('A1:W1')
        header = ws['A1']
        header.value = f'Загруженность преподавателя: {teacher_name} ({semester} семестр)'
        header.font = header_font
        header.alignment = center
        header.border = borders['thick']

        current_row = 2

        for subject, data in report['subjects'].items():
            ws.merge_cells(f'A{current_row}:W{current_row}')
            subject_cell = ws[f'A{current_row}']
            subject_cell.value = f'Предмет: {subject}'
            subject_cell.font = header_font
            subject_cell.border = borders['thick']
            current_row += 1

            for group, group_data in data['groups'].items():
                # Заголовок группы
                group_cell = ws.cell(row=current_row, column=1, value=f'Группа: {group}')
                group_cell.font = header_font
                group_cell.border = borders['thick']

                # Заголовки недель
                for i in range(21):
                    week_cell = ws.cell(row=current_row, column=i + 2, value=f'Неделя {i + 1}')
                    week_cell.font = header_font
                    week_cell.border = borders['thick']
                    week_cell.alignment = center

                total_header = ws.cell(row=current_row, column=23, value='ИТОГО')
                total_header.font = header_font
                total_header.border = borders['thick']
                total_header.alignment = center
                current_row += 1

                # Типы занятий
                for lesson_type, label in [
                    ('lecture', 'Лекции'),
                    ('practice', 'Практики'),
                    ('laboratory', 'Лабораторные')
                ]:
                    type_cell = ws.cell(row=current_row, column=1, value=label)
                    type_cell.font = normal_font
                    type_cell.border = borders['thin']

                    total = 0
                    for week in range(1, 22):
                        week_data = group_data['by_week'][str(week)]
                        cell = ws.cell(row=current_row, column=week + 1)

                        # Проверяем наличие экзамена/зачета
                        if week_data['exam_type']:
                            cell.value = week_data['exam_type']
                            cell.fill = exam_fill
                        else:
                            hours = week_data[lesson_type]
                            cell.value = hours if hours > 0 else ''
                            if hours > 0:
                                total += hours

                        cell.font = normal_font
                        cell.border = borders['thin']
                        cell.alignment = center

                    total_cell = ws.cell(row=current_row, column=23, value=total)
                    total_cell.font = normal_font
                    total_cell.border = borders['thin']
                    total_cell.alignment = center
                    current_row += 1

                current_row += 1
            current_row += 1

        # Форматирование
        ws.column_dimensions['A'].width = 25
        for col_idx in range(23):
            ws.column_dimensions[chr(ord('B') + col_idx)].width = 12
        ws.column_dimensions['W'].width = 12

        # Настройки печати
        ws.print_area = f'A1:W{current_row - 1}'
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = '1:1'

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

    @staticmethod
    def export_teacher_load_excel_from_report(report: Dict) -> BytesIO:
        """Exports an Excel file from a pre-generated report."""
        wb = Workbook()
        ws = wb.active

        # Стили
        header_font = Font(name='Times New Roman', bold=True)
        normal_font = Font(name='Times New Roman')
        borders = {
            'thick': Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            ),
            'thin': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        center = Alignment(horizontal='center', vertical='center')
        exam_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

        # Заголовок
        ws.merge_cells('A1:W1')
        header = ws['A1']
        header.value = f'Загруженность преподавателя: {report["teacher_name"]} ({report["semester"]} семестр)'
        header.font = header_font
        header.alignment = center
        header.border = borders['thick']

        current_row = 2

        for subject, data in report['subjects'].items():
            ws.merge_cells(f'A{current_row}:W{current_row}')
            subject_cell = ws[f'A{current_row}']
            subject_cell.value = f'Предмет: {subject}'
            subject_cell.font = header_font
            subject_cell.border = borders['thick']
            current_row += 1

            for group, group_data in data['groups'].items():
                # Заголовок группы
                group_cell = ws.cell(row=current_row, column=1, value=f'Группа: {group}')
                group_cell.font = header_font
                group_cell.border = borders['thick']

                # Заголовки недель
                for i in range(21):
                    week_cell = ws.cell(row=current_row, column=i + 2, value=f'Неделя {i + 1}')
                    week_cell.font = header_font
                    week_cell.border = borders['thick']
                    week_cell.alignment = center

                total_header = ws.cell(row=current_row, column=23, value='ИТОГО')
                total_header.font = header_font
                total_header.border = borders['thick']
                total_header.alignment = center
                current_row += 1

                # Типы занятий
                for lesson_type, label in [
                    ('lecture', 'Лекции'),
                    ('practice', 'Практики'),
                    ('laboratory', 'Лабораторные')
                ]:
                    type_cell = ws.cell(row=current_row, column=1, value=label)
                    type_cell.font = normal_font
                    type_cell.border = borders['thin']

                    total = 0
                    for week in range(1, 22):
                        week_data = group_data['by_week'][str(week)]
                        cell = ws.cell(row=current_row, column=week + 1)

                        # Проверяем наличие экзамена/зачета
                        if week_data['exam_type']:
                            cell.value = week_data['exam_type']
                            cell.fill = exam_fill
                        else:
                            hours = week_data[lesson_type]
                            cell.value = hours if hours > 0 else ''
                            if hours > 0:
                                total += hours

                        cell.font = normal_font
                        cell.border = borders['thin']
                        cell.alignment = center

                    total_cell = ws.cell(row=current_row, column=23, value=total)
                    total_cell.font = normal_font
                    total_cell.border = borders['thin']
                    total_cell.alignment = center
                    current_row += 1

                current_row += 1
            current_row += 1

        # Форматирование
        ws.column_dimensions['A'].width = 25
        for col_idx in range(22):
            ws.column_dimensions[chr(ord('B') + col_idx)].width = 12
        ws.column_dimensions['W'].width = 12

        # Настройки печати
        ws.print_area = f'A1:W{current_row - 1}'
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = '1:1'

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

    @staticmethod
    def generate_schedule_list(teacher_name: str, start_date: str, end_date: str) -> BytesIO:
        """
        Генерирует отчет со списком занятий преподавателя за указанный период
        с группировкой занятий при совпадении параметров
        """
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Получаем все занятия за период
        schedule_entries = Schedule.query.filter(
            Schedule.teacher_name == teacher_name,
            Schedule.date >= start_dt,
            Schedule.date <= end_dt
        ).order_by(Schedule.date, Schedule.time_start).all()

        # Группируем занятия с одинаковыми параметрами
        grouped_entries = []
        temp_dict = {}

        for entry in schedule_entries:
            # Создаем ключ для группировки
            # Включаем подгруппу в ключ, чтобы различать занятия по подгруппам
            key = (
                entry.date,
                entry.time_start,
                entry.time_end,
                entry.subject,
                entry.auditory,
                entry.lesson_type,
                entry.subgroup  # Добавляем подгруппу в ключ
            )

            # Группируем занятия с совпадающими параметрами
            if key in temp_dict:
                # Добавляем группу к существующей записи
                temp_dict[key]['groups'].append(entry.group_name)
            else:
                # Создаем новую запись
                temp_dict[key] = {
                    'entry': entry,
                    'groups': [entry.group_name]
                }

        # Преобразуем сгруппированные записи в список
        for key, value in temp_dict.items():
            entry = value['entry']
            entry.grouped_groups = ', '.join(sorted(value['groups']))
            grouped_entries.append(entry)

        # Сортируем все записи
        grouped_entries.sort(key=lambda x: (x.date, x.time_start, x.time_end, x.subject))

        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание занятий"

        # Стили
        header_font = Font(name='Times New Roman', bold=True)
        normal_font = Font(name='Times New Roman')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок
        ws.merge_cells('A1:G1')
        header = ws['A1']
        header.value = f'Расписание занятий преподавателя: {teacher_name}'
        header.font = header_font
        header.alignment = Alignment(horizontal='center')

        # Заголовки колонок
        headers = ['Дата', 'Время', 'Тип занятия', 'Предмет', 'Группа(ы)', 'Аудитория', 'Подгруппа']
        for col, header_text in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header_text
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Заполняем данные
        current_row = 3
        current_date = None

        for entry in grouped_entries:
            # Форматируем дату
            date_str = entry.date.strftime('%d.%m.%Y')

            # Добавляем пустую строку между датами
            if current_date and current_date != date_str:
                current_row += 1
            current_date = date_str

            # Определяем группы
            groups = getattr(entry, 'grouped_groups', entry.group_name)

            # Записываем данные о занятии
            row_data = [
                date_str,
                f"{entry.time_start} - {entry.time_end}",
                entry.lesson_type,
                entry.subject,
                groups,
                entry.auditory or '',
                f"Подгруппа {entry.subgroup}" if entry.subgroup != 0 else ""
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col > 2 else 'center')

            current_row += 1

        # Настраиваем ширину колонок
        column_widths = [12, 15, 15, 40, 25, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Настройки печати
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = '1:2'

        # Добавляем информацию о периоде под таблицей
        footer_row = current_row + 2
        ws.merge_cells(f'A{footer_row}:G{footer_row}')
        footer = ws[f'A{footer_row}']
        footer.value = f'Период: с {start_date} по {end_date}'
        footer.font = normal_font
        footer.alignment = Alignment(horizontal='left')

        # Сохраняем файл
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    @staticmethod
    def get_multiple_teachers_load(semester: int, teachers: List[str]) -> Dict:
        """
        Получает нагрузки для списка преподавателей за один запрос к базе.
        Возвращает словарь формата:
        {
            'Имя Преподавателя': {
                'teacher_name': ...,
                'semester': ...,
                'subjects': {
                    'предмет': {
                        'groups': {
                            'группа': {
                                'faculty': ...,
                                'total_hours': ...,
                                'by_type': {...},
                                'by_week': {...}
                            }
                        },
                        'total_hours': ...,
                        'by_type': {...}
                    }
                }
            }
        }
        """
        # Инициализируем структуру
        reports = {}
        for teacher in teachers:
            reports[teacher] = {
                'teacher_name': teacher,
                'semester': semester,
                'subjects': {}
            }

        # Один большой запрос
        lessons = (db.session.query(Schedule)
                   .filter(Schedule.semester == semester,
                           Schedule.teacher_name.in_(teachers))
                   .order_by(Schedule.week_number, Schedule.date)
                   .all())

        # Обработка данных
        for lesson in lessons:
            teacher_name = lesson.teacher_name
            subject = lesson.subject
            group = lesson.group_name
            faculty = lesson.faculty
            week_num = lesson.week_number
            original_type = lesson.lesson_type
            lesson_type = ReportService.LESSON_TYPES.get(original_type, 'other')
            hours = 2  # Допустим 2 часа на пару

            teacher_report = reports[teacher_name]
            if subject not in teacher_report['subjects']:
                teacher_report['subjects'][subject] = {
                    'groups': {},
                    'total_hours': 0,
                    'by_type': {
                        'lecture': 0,
                        'practice': 0,
                        'laboratory': 0,
                        'other': 0,
                        'exam': []
                    }
                }
            subj_data = teacher_report['subjects'][subject]

            if group not in subj_data['groups']:
                subj_data['groups'][group] = {
                    'faculty': faculty,
                    'total_hours': 0,
                    'by_type': {
                        'lecture': 0,
                        'practice': 0,
                        'laboratory': 0,
                        'other': 0
                    },
                    'by_week': {str(w): {
                        'lecture': 0,
                        'practice': 0,
                        'laboratory': 0,
                        'other': 0,
                        'exam_type': None,
                        'hours': {}
                    } for w in range(1, 20)}
                }
            group_data = subj_data['groups'][group]
            week_data = group_data['by_week'][str(week_num)]

            # Записываем часы
            if lesson_type in ['lecture', 'practice', 'laboratory']:
                week_data[lesson_type] += hours
                group_data['total_hours'] += hours
                group_data['by_type'][lesson_type] += hours
                subj_data['total_hours'] += hours
                subj_data['by_type'][lesson_type] += hours

            # Экзамены/зачеты
            if original_type in ReportService.EXAM_DISPLAY:
                week_data['exam_type'] = ReportService.EXAM_DISPLAY[original_type]
                subj_data['by_type']['exam'].append({
                    'type': ReportService.EXAM_DISPLAY[original_type],
                    'week': week_num,
                    'group': group
                })

        return reports

    @staticmethod
    def get_teachers_total_load(semester: int) -> List[Dict]:
        """Получает общую нагрузку всех преподавателей за семестр"""
        load_data = db.session.query(
            Schedule.teacher_name,
            func.count(Schedule.id).label('lessons_count'),
            func.count(distinct(Schedule.subject)).label('subjects_count'),
            func.count(distinct(Schedule.week_number)).label('weeks_count')
        ).filter(
            Schedule.semester == semester,
            Schedule.teacher_name != ''
        ).group_by(Schedule.teacher_name).all()

        result = []
        for data in load_data:
            hours = data.lessons_count * 2  # 1 пара = 2 академических часа
            result.append({
                'teacher_name': data.teacher_name,
                'total_hours': hours,
                'subjects_count': data.subjects_count,
                'weeks_count': data.weeks_count
            })

        return sorted(result, key=lambda x: x['total_hours'], reverse=True)